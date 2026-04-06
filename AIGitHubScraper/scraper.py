import os
import smtplib
import requests
from datetime import datetime, timedelta, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()

GITHUB_TOKEN  = os.getenv("GITHUB_TOKEN")
GMAIL_USER    = os.getenv("GMAIL_USER")     # your Gmail address
GMAIL_APP_PW  = os.getenv("GMAIL_APP_PW")  # Gmail App Password (not your login password)
EMAIL_TO      = os.getenv("EMAIL_TO", GMAIL_USER)  # defaults to sender if not set
OUTPUT_DIR    = os.path.join(os.path.dirname(__file__), "output")

HEADERS = {
    "Accept": "application/vnd.github+json",
    "X-GitHub-Api-Version": "2022-11-28",
}
if GITHUB_TOKEN:
    HEADERS["Authorization"] = f"Bearer {GITHUB_TOKEN}"

AI_TOPICS = [
    "large-language-model",
    "llm",
    "generative-ai",
    "ai-agent",
    "machine-learning",
    "deep-learning",
    "diffusion-model",
    "transformer",
    "rag",
    "multimodal",
]

AI_KEYWORDS = [
    "mcp server",
    "vibe coding",
    "ai coding",
    "reasoning model",
]

KNOWN_AFFILIATIONS = {
    "google": "Google",
    "deepmind": "Google DeepMind",
    "meta": "Meta",
    "facebook": "Meta",
    "microsoft": "Microsoft",
    "openai": "OpenAI",
    "anthropic": "Anthropic",
    "apple": "Apple",
    "amazon": "Amazon / AWS",
    "aws": "Amazon / AWS",
    "nvidia": "NVIDIA",
    "hugging face": "Hugging Face",
    "huggingface": "Hugging Face",
    "stanford": "Stanford University",
    "mit": "MIT",
    "berkeley": "UC Berkeley",
    "cmu": "Carnegie Mellon",
    "carnegie mellon": "Carnegie Mellon",
    "oxford": "University of Oxford",
    "cambridge": "University of Cambridge",
    "eth zurich": "ETH Zurich",
    "tsinghua": "Tsinghua University",
    "bytedance": "ByteDance",
    "alibaba": "Alibaba",
    "baidu": "Baidu",
    "tencent": "Tencent",
    "mistral": "Mistral AI",
    "cohere": "Cohere",
    "inflection": "Inflection AI",
}

MIN_STARS = 200
RESULTS_PER_QUERY = 10
TRENDING_DAYS = 7


# ── GitHub API helpers ────────────────────────────────────────────────────────

def search_topic(topic: str, since: str) -> list[dict]:
    params = {
        "q": f"topic:{topic} created:>{since} stars:>={MIN_STARS}",
        "sort": "stars",
        "order": "desc",
        "per_page": RESULTS_PER_QUERY,
    }
    r = requests.get("https://api.github.com/search/repositories",
                     headers=HEADERS, params=params, timeout=15)
    r.raise_for_status()
    items = r.json().get("items", [])
    for item in items:
        item.setdefault("_matched_topics", [])
        item["_matched_topics"].append(topic)
    return items


def search_keyword(keyword: str, since: str) -> list[dict]:
    params = {
        "q": f"{keyword} in:name,description,readme created:>{since} stars:>={MIN_STARS}",
        "sort": "stars",
        "order": "desc",
        "per_page": RESULTS_PER_QUERY,
    }
    r = requests.get("https://api.github.com/search/repositories",
                     headers=HEADERS, params=params, timeout=15)
    r.raise_for_status()
    items = r.json().get("items", [])
    for item in items:
        item.setdefault("_matched_topics", [])
        item["_matched_topics"].append(keyword)
    return items


def get_user_profile(username: str) -> dict:
    r = requests.get(f"https://api.github.com/users/{username}",
                     headers=HEADERS, timeout=15)
    if r.status_code != 200:
        return {}
    return r.json()


# ── Enrichment helpers ────────────────────────────────────────────────────────

def infer_affiliation(profile: dict) -> str:
    company = (profile.get("company") or "").strip().lstrip("@")
    bio = (profile.get("bio") or "").strip()
    text = f"{company} {bio}".lower()

    for key, label in KNOWN_AFFILIATIONS.items():
        if key in text:
            return label

    if company:
        return company
    return "Independent / Unknown"


def infer_trend_reason(repo: dict) -> str:
    created = datetime.fromisoformat(repo["created_at"].replace("Z", "+00:00"))
    days_old = (datetime.now(timezone.utc) - created).days
    stars = repo["stargazers_count"]
    topics = repo.get("_matched_topics", [])

    parts = []

    if days_old <= 3:
        parts.append("Just launched")
    elif days_old <= 7:
        parts.append("New release")
    elif days_old <= 14:
        parts.append("Recent project gaining traction")

    if stars >= 5000:
        parts.append(f"viral ({stars:,} stars)")
    elif stars >= 1000:
        parts.append("rapidly growing")

    if topics:
        clean = [t.replace("-", " ") for t in topics[:3]]
        parts.append(f"trending in: {', '.join(clean)}")

    return "; ".join(parts) if parts else "Highly starred AI project"


# ── Dedup + rank ──────────────────────────────────────────────────────────────

def dedupe(repos: list[dict]) -> list[dict]:
    seen: dict[int, dict] = {}
    for repo in repos:
        rid = repo["id"]
        if rid not in seen:
            seen[rid] = repo
        else:
            # merge matched topics
            seen[rid].setdefault("_matched_topics", [])
            seen[rid]["_matched_topics"] += repo.get("_matched_topics", [])
    return sorted(seen.values(), key=lambda r: r["stargazers_count"], reverse=True)


# ── Excel output ──────────────────────────────────────────────────────────────

COLUMNS = [
    ("Repo Name",          28),
    ("Author",             18),
    ("Author Affiliation", 28),
    ("Description",        55),
    ("Trend Reason",       45),
    ("Stars",              10),
    ("Topic",              30),
]

HEADER_FILL   = PatternFill("solid", fgColor="1F3864")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL      = PatternFill("solid", fgColor="EEF2FF")
HYPERLINK_FONT = Font(color="1155CC", underline="single")


def write_excel(rows: list[dict], path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AI Trending Repos"
    ws.freeze_panes = "A2"

    # Header row
    for col_idx, (header, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22

    for i, row in enumerate(rows, 2):
        fill = ALT_FILL if i % 2 == 0 else None

        name_cell = ws.cell(row=i, column=1, value=row["repo_name"])
        name_cell.hyperlink = row["url"]
        name_cell.font = HYPERLINK_FONT

        ws.cell(row=i, column=2, value=row["author"])
        ws.cell(row=i, column=3, value=row["affiliation"])
        ws.cell(row=i, column=4, value=row["description"])
        ws.cell(row=i, column=5, value=row["trend_reason"])
        ws.cell(row=i, column=6, value=row["stars"])
        ws.cell(row=i, column=7, value=row["topic"])

        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=i, column=col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in (4, 5)))
            if fill:
                cell.fill = fill

        ws.row_dimensions[i].height = 40

    wb.save(path)


# ── Email ─────────────────────────────────────────────────────────────────────

AFFILIATION_COLORS = {
    "Google": "#4285F4", "Google DeepMind": "#4285F4",
    "Meta": "#0866FF", "Microsoft": "#00A4EF",
    "OpenAI": "#10A37F", "Anthropic": "#C97B3B",
    "NVIDIA": "#76B900", "Hugging Face": "#FFD21E",
    "MIT": "#A31F34", "Stanford University": "#8C1515",
    "UC Berkeley": "#003262", "Carnegie Mellon": "#C41230",
}

def _affiliation_badge(affiliation: str) -> str:
    color = AFFILIATION_COLORS.get(affiliation, "#6B7280")
    return (
        f'<span style="display:inline-block;padding:2px 10px;border-radius:12px;'
        f'background:{color};color:#fff;font-size:11px;font-weight:600;'
        f'letter-spacing:0.3px">{affiliation}</span>'
    )

def _star_bar(stars: int, max_stars: int) -> str:
    pct = min(100, int(stars / max_stars * 100)) if max_stars else 0
    return (
        f'<div style="display:flex;align-items:center;gap:8px">'
        f'<div style="flex:1;height:6px;background:#E5E7EB;border-radius:3px">'
        f'<div style="width:{pct}%;height:100%;background:#FBBF24;border-radius:3px"></div></div>'
        f'<span style="font-size:13px;font-weight:700;color:#1F2937">⭐ {stars:,}</span>'
        f'</div>'
    )

def build_email_html(rows: list[dict], date_str: str) -> str:
    max_stars = rows[0]["stars"] if rows else 1

    cards_html = ""
    for i, row in enumerate(rows, 1):
        cards_html += f"""
        <div style="background:#fff;border:1px solid #E5E7EB;border-radius:12px;
                    padding:24px;margin-bottom:20px;box-shadow:0 1px 4px rgba(0,0,0,.06)">
          <div style="display:flex;justify-content:space-between;align-items:flex-start;
                      flex-wrap:wrap;gap:8px;margin-bottom:12px">
            <div>
              <span style="color:#9CA3AF;font-size:12px;font-weight:600;
                           text-transform:uppercase;letter-spacing:1px">#{i}</span>
              <a href="{row['url']}"
                 style="display:block;font-size:20px;font-weight:700;color:#111827;
                        text-decoration:none;margin-top:2px;line-height:1.3">
                {row['repo_name']}
              </a>
              <span style="font-size:13px;color:#6B7280">by&nbsp;
                <strong style="color:#374151">{row['author']}</strong>
              </span>
            </div>
            <div style="text-align:right">
              {_affiliation_badge(row['affiliation'])}
            </div>
          </div>

          <p style="margin:0 0 14px;font-size:14px;color:#374151;line-height:1.6">
            {row['description'] or '<em style="color:#9CA3AF">No description provided.</em>'}
          </p>

          <div style="background:#F9FAFB;border-radius:8px;padding:10px 14px;
                      margin-bottom:14px;font-size:13px;color:#6B7280">
            <span style="font-weight:600;color:#374151">Why it's trending:&nbsp;</span>
            {row['trend_reason']}
          </div>

          {_star_bar(row['stars'], max_stars)}
        </div>"""

    return f"""<!DOCTYPE html>
<html lang="en">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>AI GitHub Trending — {date_str}</title></head>
<body style="margin:0;padding:0;background:#F3F4F6;font-family:-apple-system,BlinkMacSystemFont,
             'Segoe UI',Roboto,Helvetica,Arial,sans-serif">
  <div style="max-width:680px;margin:32px auto;padding:0 16px">

    <!-- Header -->
    <div style="background:linear-gradient(135deg,#1E1B4B 0%,#312E81 50%,#4338CA 100%);
                border-radius:16px;padding:36px 32px;margin-bottom:28px;text-align:center">
      <div style="font-size:28px;margin-bottom:8px">🤖</div>
      <h1 style="margin:0;color:#fff;font-size:26px;font-weight:800;letter-spacing:-0.5px">
        AI GitHub Trending
      </h1>
      <p style="margin:8px 0 0;color:#A5B4FC;font-size:14px">{date_str} &nbsp;·&nbsp;
        Top {len(rows)} repositories this week</p>
    </div>

    <!-- Cards -->
    {cards_html}

    <!-- Footer -->
    <div style="text-align:center;padding:24px 0 40px;font-size:12px;color:#9CA3AF">
      Generated automatically by AIGitHubScraper &nbsp;·&nbsp;
      <a href="https://github.com/Boxxxi/AI/tree/main/AIGitHubScraper/output"
         style="color:#6366F1;text-decoration:none">View all reports</a>
    </div>

  </div>
</body>
</html>"""


def send_email(html: str, date_str: str):
    if not GMAIL_USER or not GMAIL_APP_PW:
        print("  GMAIL_USER / GMAIL_APP_PW not set — skipping email.")
        return

    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"🤖 AI GitHub Trending — {date_str}"
    msg["From"]    = GMAIL_USER
    msg["To"]      = EMAIL_TO
    msg.attach(MIMEText(html, "html"))

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(GMAIL_USER, GMAIL_APP_PW)
        smtp.sendmail(GMAIL_USER, EMAIL_TO, msg.as_string())

    print(f"  Email sent to {EMAIL_TO}")


# ── Main ──────────────────────────────────────────────────────────────────────

def run():
    since = (datetime.now(timezone.utc) - timedelta(days=TRENDING_DAYS)).strftime("%Y-%m-%d")
    print(f"Fetching AI repos trending since {since} ...\n")

    raw: list[dict] = []

    for topic in AI_TOPICS:
        try:
            results = search_topic(topic, since)
            raw.extend(results)
            print(f"  topic:{topic:30s} -> {len(results)} results")
        except requests.HTTPError as e:
            print(f"  topic:{topic} ERROR: {e}")

    for kw in AI_KEYWORDS:
        try:
            results = search_keyword(kw, since)
            raw.extend(results)
            print(f"  keyword:'{kw}'{' ' * max(0, 22 - len(kw))} -> {len(results)} results")
        except requests.HTTPError as e:
            print(f"  keyword:'{kw}' ERROR: {e}")

    ranked = dedupe(raw)[:50]
    print(f"\nEnriching {len(ranked)} repos with author profiles ...\n")

    rows = []
    for repo in ranked:
        username = repo["owner"]["login"]
        profile = get_user_profile(username)
        affiliation = infer_affiliation(profile)
        trend_reason = infer_trend_reason(repo)

        matched = list(dict.fromkeys(repo.get("_matched_topics", [])))  # deduped, ordered
        rows.append({
            "repo_name":   repo["name"],
            "author":      username,
            "affiliation": affiliation,
            "description": repo.get("description") or "",
            "trend_reason": trend_reason,
            "stars":       repo["stargazers_count"],
            "topic":       ", ".join(matched),
            "url":         repo["html_url"],
        })

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    today = datetime.now().strftime("%Y-%m-%d")
    out_path = os.path.join(OUTPUT_DIR, f"AIGitHubScraper_{today}.xlsx")
    write_excel(rows, out_path)
    print(f"Saved {len(rows)} repos -> {out_path}\n")

    print("Sending email digest ...")
    today_pretty = datetime.now().strftime("%B %d, %Y")
    html = build_email_html(rows, today_pretty)
    send_email(html, today_pretty)

    print(f"\n{'#':<4} {'Stars':>7}  {'Affiliation':25}  Repo")
    print("-" * 70)
    for i, row in enumerate(rows[:20], 1):
        print(f"{i:<4} {row['stars']:>7,}  {row['affiliation']:25}  {row['author']}/{row['repo_name']}")


if __name__ == "__main__":
    run()
